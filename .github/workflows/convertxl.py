import os
import json
from openpyxl import load_workbook


def create_output(base_path, sheet_name):
    path = os.path.join(base_path, sheet_name)
    os.makedirs(path, exist_ok=True)
    return path


def get_sheet_data(ws):
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], []

    headers = list(rows[0])
    data_rows = rows[1:]

    return headers, data_rows


def write_txt(headers, data_rows, columns, output_dir, sheet_name):
    # Enforce exactly one column
    if len(columns) != 1:
        raise ValueError(
            f"TXT format in sheet '{sheet_name}' must have exactly ONE column"
        )

    col = columns[0]

    col_index = {col_name: idx for idx, col_name in enumerate(headers)}

    if col not in col_index:
        raise ValueError(f"Column '{col}' not found in sheet '{sheet_name}'")

    idx = col_index[col]

    # File name = sheet name
    filepath = os.path.join(output_dir, f"{sheet_name}.txt")

    with open(filepath, "w", encoding="utf-8") as f:
        for row in data_rows:
            value = row[idx]
            if value is not None and value != "":
                f.write(f"{value}\n")


def write_csv(headers, data_rows, columns, output_dir, sheet_name):
    filepath = os.path.join(output_dir, f"{sheet_name}.csv")

    col_index = {col: idx for idx, col in enumerate(headers)}
    selected_indexes = [col_index[col] for col in columns if col in col_index]

    with open(filepath, "w", encoding="utf-8") as f:
        # header
        f.write(",".join(columns) + "\n")

        # rows
        for row in data_rows:
            values = [
                str(row[i]) if row[i] is not None else ""
                for i in selected_indexes
            ]
            f.write(",".join(values) + "\n")


def write_json(headers, data_rows, columns, output_dir, sheet_name):
    filepath = os.path.join(output_dir, f"{sheet_name}.json")

    col_index = {col: idx for idx, col in enumerate(headers)}

    result = []
    for row in data_rows:
        obj = {}
        for col in columns:
            if col in col_index:
                value = row[col_index[col]]
                obj[col] = value
        result.append(obj)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)


def process_excel():
    excel_path = "lists-of-things.xlsx"
    config_path = ".github/workflows/config.json"
    output_folder = "."

    # Load config
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    # Load workbook
    wb = load_workbook(excel_path, data_only=True)

    for sheet_name, rules in config.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in Excel")
            continue

        ws = wb[sheet_name]
        headers, data_rows = get_sheet_data(ws)

        if not headers:
            print(f"Sheet '{sheet_name}' is empty")
            continue

        output_dir = create_output(output_folder, sheet_name)

        for fmt, columns in rules.items():
            missing_cols = [col for col in columns if col not in headers]
            if missing_cols:
                print(f"Missing columns in {sheet_name}: {missing_cols}")
                continue

            try:
                if fmt == "txt":
                    write_txt(headers, data_rows, columns, output_dir, sheet_name)

                elif fmt == "csv":
                    write_csv(headers, data_rows, columns, output_dir, sheet_name)

                elif fmt == "json":
                    write_json(headers, data_rows, columns, output_dir, sheet_name)

                else:
                    print(f"Unsupported format '{fmt}' in sheet '{sheet_name}'")

            except ValueError as e:
                print(f"{e}")


if __name__ == "__main__":
    process_excel()